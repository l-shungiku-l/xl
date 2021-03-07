import os
import glob
import sys

# openpyxl is required
import openpyxl as px


def recursively_get_file(dirname: str) -> list:
    file_list = glob.glob(dirname + "\\**\\*.xlsx", recursive=True)
    return [f for f in file_list]


class SGBook:
    def __init__(self, filename: str) -> None:
        self.filename = filename
        self.workbook = px.load_workbook(filename, keep_vba=True, data_only=True)

    def sheet(self, target_sheetname: str):
        sheet_list = self.get_sheet_list()
        if target_sheetname in sheet_list:
            return SGSheet(self, target_sheetname)
        else:
            print("ERROR: " + target_sheetname + " is not found in " + self.filename)
            print("all sheets in the book are below")
            print(sheet_list)
            sys.exit()

    def get_sheet_list(self):
        return self.workbook.sheetnames

    def save_and_close(self):
        self.workbook.save(self.filename)
        self.workbook.close()
        print(self.filename + " has been closed")

    def close(self):
        self.workbook.close()
        print(self.filename + " has been closed")


class SGSheet:
    def __init__(self, sgbook: SGBook, target_sheetname: str) -> None:
        self.sgbook = sgbook
        self.sheet = self.sgbook.workbook[target_sheetname]
        self.cell = self.sheet.cell

    def get_basename(self, y, x) -> str:
        return os.path.basename(self.cell(y, x).value)

    def get_dirname(self, y, x) -> str:
        return os.path.dirname(self.cell(y, x).value)

    def get_extnsion(self, y, x) -> str:
        return os.path.splitext(self.cell(y, x).value)[1]

    def remove_extension(self, y, x) -> str:
        return os.path.splitext(self.cell(y, x).value)[0]

    def get_row(self, y, x) -> list:
        row = [
            c[0].value for c in self.sheet.iter_cols(min_row=y, max_row=y, min_col=x)
        ]
        self._remove_last_none(row)
        return row

    def get_col(self, y, x) -> list:
        col = [
            c[0].value for c in self.sheet.iter_rows(min_row=y, min_col=x, max_col=x)
        ]
        self._remove_last_none(col)
        return col

    def get_all_cells(self) -> list:
        list_2d = []
        for row in self.sheet.iter_rows():
            list_2d.append([c.value for c in row])
        return list_2d

    def write_row(self, list_1d: list, start_row: int, start_col: int) -> None:
        for x, c in enumerate(list_1d):
            self.cell(start_row, start_col + x).value = c

    def write_col(self, list_1d: list, start_row: int, start_col: int) -> None:
        for y, c in enumerate(list_1d):
            self.cell(start_row + y, start_col).value = c

    def write_list_2d(self, list_2d, start_row: int, start_col: int) -> None:
        for y, row in enumerate(list_2d):
            for x, cell in enumerate(row):
                self.cell(start_row + y, start_col + x).value = list_2d[y][x]

    def _remove_last_none(self, list_1d):
        if list_1d[-1] == None:
            del list_1d[-1]
            self._remove_last_none(list_1d)
